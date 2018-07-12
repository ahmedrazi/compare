import openpyxl
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
wb = load_workbook('data.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

# Get data from device_info spreasheet
ip1 = sheet['A1'].value
user1 = sheet['B1'].value
password1 = sheet['C1'].value
device1type = sheet['D1'].value
for i in range(1, 3):

    tree = ET.parse('data.xml')
    root= tree.getroot()
    #print(root.findall([sheet.cell(row=i, column=2).value]))
    print(root.findall("./Age))






